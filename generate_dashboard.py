# -*- coding: utf-8 -*-
"""
서울동부지사 정비사업 뉴스+고시공고 대시보드 자동 생성기 (v12)
- 뉴스: 네이버 검색 API (최근 30일, 7개 구)
- 고시공고: 구청별 공식 고시공고 게시판 바로가기 탭 제공
- 최근 실거래: 국토부 실거래가 API로 구별 아파트 매매/전세/월세 (계약일 기준 최근 7일)
- 추진현황: 서울시 도시정비사업 통계 분기 엑셀(data/*.xlsx) 기반 구역별 단계 진행바
- 결과를 docs/index.html 에 저장 (GitHub Pages 배포)

필요 라이브러리: requests
"""

import os
import re
import html
import json
import requests
from datetime import datetime, timedelta, timezone
from email.utils import parsedate_to_datetime

# ──────────────────────────────────────────────
# 공통 설정
# ──────────────────────────────────────────────
NAVER_CLIENT_ID = os.environ.get("NAVER_CLIENT_ID", "")
NAVER_CLIENT_SECRET = os.environ.get("NAVER_CLIENT_SECRET", "")

# 대시보드 상단 공지줄 (비우면 표시 안 됨). 내용 수정 후 커밋하면 다음 갱신에 반영
UPDATE_NOTICE = "🆕 2026-07-10 · '최근 실거래' 탭 신설 — 구별 아파트 매매/전세/월세 최근 7일 계약분 제공"

DISTRICTS = ["성동구", "광진구", "동대문구", "중랑구", "도봉구", "노원구", "강북구"]
KEYWORDS = ["정비사업", "재개발", "재건축", "재정비", "모아타운", "신속통합기획", "공공주택 복합"]

DAYS_BACK = 30
MAX_PER_QUERY = 30
MAX_PER_DISTRICT = 15
OUTPUT_PATH = os.path.join("docs", "index.html")

RELEVANCE_WORDS = KEYWORDS + ["조합", "관리처분", "사업시행", "안전진단", "이주", "착공",
                              "분양", "시공사", "정비구역", "조합설립", "리모델링"]

# 고시 제목에 이 단어가 있으면 정비사업 관련 고시로 채택
NOTICE_KEYWORDS = ["정비", "재개발", "재건축", "모아타운", "신속통합", "리모델링",
                   "관리처분", "사업시행", "조합", "정비구역", "도시계획", "지구단위"]

KST = timezone(timedelta(hours=9))

# ──────────────────────────────────────────────
# 고시공고: 구청 공식 게시판 바로가기 (링크는 2026-07 확인)
# ──────────────────────────────────────────────
NOTICE_BOARDS = {
    "성동구": {
        "name": "고시공고/입법예고 게시판",
        "url": "https://www.sd.go.kr/main/selectBbsNttList.do?bbsNo=184&key=1473",
    },
    "광진구": {
        "name": "고시공고 게시판",
        "url": "https://www.gwangjin.go.kr/portal/bbs/B0000003/list.do?menuNo=200192",
    },
    "동대문구": {
        "name": "고시공고 게시판",
        "url": "https://www.ddm.go.kr/www/selectEminwonWebList.do?key=3291&searchNotAncmtSeCode=01,02,04,05,06,07",
    },
    "중랑구": {
        "name": "공고/고시 게시판",
        "url": "https://www.jungnang.go.kr/portal/bbs/list/B0000117.do?menuNo=200475",
    },
    "도봉구": {
        "name": "고시/공고 게시판",
        "url": "https://www.dobong.go.kr/WDB_DEV/gosigong_go/",
    },
    "노원구": {
        "name": "고시공고 게시판",
        "url": "https://www.nowon.kr/www/user/bbs/BD_selectBbsList.do?q_bbsCode=1003&q_clCode=0&q_estnColumn1=11&q_ntceSiteCode=11",
    },
    "강북구": {
        "name": "고시공고 게시판",
        "url": "https://www.gangbuk.go.kr/portal/bbs/B0000245/list.do?menuNo=200082",
    },
}


# ──────────────────────────────────────────────
# 최근 실거래 (국토교통부 실거래가 공개 API)
# ──────────────────────────────────────────────
MOLIT_KEY = os.environ.get("DATA_GO_KR_KEY", "")  # 공공데이터포털 인증키(Decoding)

LAWD_CD = {  # 법정동 시군구코드
    "성동구": "11200", "광진구": "11215", "동대문구": "11230", "중랑구": "11260",
    "도봉구": "11320", "노원구": "11350", "강북구": "11305",
}

DEAL_DAYS_BACK = 7          # 계약일 기준 최근 며칠
MAX_DEALS_PER_DISTRICT = 40 # 구별 표시 상한

TRADE_URL = "https://apis.data.go.kr/1613000/RTMSDataSvcAptTradeDev/getRTMSDataSvcAptTradeDev"
RENT_URL = "https://apis.data.go.kr/1613000/RTMSDataSvcAptRent/getRTMSDataSvcAptRent"

# 카카오맵: REST 키는 좌표 변환(서버측), JS 키는 지도 표시(페이지측)
KAKAO_REST_KEY = os.environ.get("KAKAO_REST_KEY", "")
KAKAO_JS_KEY = os.environ.get("KAKAO_JS_KEY", "")

# 접속 게이트: GitHub Secret LOGIN_ACCOUNTS = "id1:pw1,id2:pw2" (비우면 게이트 비활성화)
LOGIN_ACCOUNTS = os.environ.get("LOGIN_ACCOUNTS", "")


def _auth_hashes() -> list:
    import hashlib
    hashes = []
    for pair in LOGIN_ACCOUNTS.split(","):
        pair = pair.strip()
        if ":" in pair:
            hashes.append(hashlib.sha256(pair.encode()).hexdigest())
    return hashes
GEO_CACHE_PATH = os.path.join("data", "geocache.json")

# 추진현황: 서울 열린데이터광장 '서울특별시 도시정비사업 통계' 분기 엑셀 (data/ 폴더에 업로드)
PROGRESS_DIR = "data"
PROGRESS_OVERRIDES = os.path.join("data", "progress_overrides.json")
STAGES = ["구역지정", "추진위", "조합설립", "건축심의", "사업시행", "관리처분", "이주", "착공", "준공"]
# 현재 단계별로 날짜를 읽을 엑셀 열 번호 (0-base)
STAGE_DATE_COL = {"구역지정": 12, "추진위": 13, "조합설립": 14, "건축심의": 15,
                  "사업시행": 17, "관리처분": 19, "이주": 20, "착공": 22}

# 지가분석: 국토부 토지 매매 실거래가 (지가변동률 조사 지원용)
LAND_URL = "https://apis.data.go.kr/1613000/RTMSDataSvcLandTrade/getRTMSDataSvcLandTrade"
LAND_MONTHS = 3               # 최근 몇 개월 사례를 볼지
MAX_LAND_ROWS = 30            # 구별 표시 사례 상한

# 토지거래허가 동향 (서울부동산정보광장, 수급 활동량 지표 — 가격정보 없음)
TOHEO_PAGE = "https://land.seoul.go.kr/land/other/contractStatus.do"
TOHEO_URL_CANDIDATES = [
    ("POST", "https://land.seoul.go.kr/land/other/contractStatusList.do"),
    ("POST", "https://land.seoul.go.kr/land/other/selectContractStatusList.do"),
    ("POST", "https://land.seoul.go.kr/land/other/selectContractStatus.do"),
    ("POST", "https://land.seoul.go.kr/land/other/contractStatus.do"),
    ("GET", "https://land.seoul.go.kr/land/other/contractStatus.do"),
]
TOHEO_PARAM_KEYS = ["cggCd", "sggCd", "sigunguCd", "guCd", "atcSggCd"]
TOHEO_ARCHIVE = os.path.join("data", "toheo_archive.json")
TOHEO_TREND_DAYS = 14
TOHEO_LIST_ROWS = 10


def _txt_any(node, tags):
    for t in tags:
        v = _txt(node, t)
        if v:
            return v
    return ""


def _median(vals):
    s = sorted(vals)
    n = len(s)
    return 0 if n == 0 else (s[n // 2] if n % 2 else (s[n // 2 - 1] + s[n // 2]) / 2)


def _toheo_parse_rows(html_text: str, district: str) -> list:
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        return []
    soup = BeautifulSoup(html_text, "html.parser")
    date_re = re.compile(r"(20\d{2})[.\-/](\d{1,2})[.\-/](\d{1,2})")
    rows = []
    for tr in soup.find_all("tr"):
        tds = [td.get_text(" ", strip=True) for td in tr.find_all("td")]
        if len(tds) < 5:
            continue
        joined = " ".join(tds)
        m = date_re.search(joined)
        if not m:
            continue
        # 표 구성: 연번 | 주소 | 지목 | 허가년월일 | 이용목적 | 이용의무종료일 | ...
        addr = max(tds, key=len)  # 가장 긴 셀 = 주소
        if len(addr) < 5:
            continue
        try:
            d = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            continue
        jimok = next((t for t in tds if t in ("대", "전", "답", "임야", "잡종지", "도로", "구거", "주차장", "창고용지", "공장용지")), "")
        purpose = next((t for t in tds if any(k in t for k in ("주거", "자기", "이용", "사업", "임대", "경영", "복지", "편익"))), "")
        rows.append({"gu": district, "addr": addr, "jimok": jimok,
                     "date": d.strftime("%Y-%m-%d"), "purpose": purpose})
    return rows


def collect_toheo(today: datetime) -> dict:
    """구별 토지거래허가 내역 수집 → 아카이브 누적 → 동향 집계"""
    # 아카이브 로드
    try:
        with open(TOHEO_ARCHIVE, encoding="utf-8") as f:
            archive = json.load(f)
    except Exception:
        archive = {}

    working = None  # 성공한 (method, url, key) 조합 기억
    ok_any = False
    for district in DISTRICTS:
        print(f"▶ {district} 토지거래허가 수집 중...")
        rows, tried = [], 0
        combos = ([working] if working else
                  [(m, u, k) for m, u in TOHEO_URL_CANDIDATES for k in TOHEO_PARAM_KEYS])
        for method, url, key in combos:
            tried += 1
            params = {key: LAWD_CD[district], "pageIndex": "1"}
            try:
                if method == "POST":
                    r = requests.post(url, data=params, headers=UA_TOHEO, timeout=12)
                else:
                    r = requests.get(url, params=params, headers=UA_TOHEO, timeout=12)
                if r.status_code != 200:
                    continue
                rows = _toheo_parse_rows(r.text, district)
            except Exception:
                continue
            if rows:
                if not working:
                    print(f"    [엔드포인트 확정] {method} {url.split('/')[-1]} · 파라미터 {key}")
                working = (method, url, key)
                break
        if rows:
            ok_any = True
            new = 0
            for x in rows:
                k = f"{x['gu']}|{x['addr']}|{x['date']}"
                if k not in archive:
                    archive[k] = x
                    new += 1
            print(f"  → 허가 {len(rows)}건 조회 (신규 {new}건 아카이브)")
        else:
            print(f"  → 수집 실패 ({tried}개 조합 시도)")

    if ok_any:
        os.makedirs(os.path.dirname(TOHEO_ARCHIVE), exist_ok=True)
        with open(TOHEO_ARCHIVE, "w", encoding="utf-8") as f:
            json.dump(archive, f, ensure_ascii=False)

    # 집계
    result = {}
    for district in DISTRICTS:
        recs = sorted((v for v in archive.values() if v["gu"] == district),
                      key=lambda x: x["date"], reverse=True)
        days = [(today - timedelta(days=i)).strftime("%Y-%m-%d") for i in range(TOHEO_TREND_DAYS - 1, -1, -1)]
        daily = [sum(1 for r in recs if r["date"] == d) for d in days]
        last7 = sum(daily[-7:])
        prev7 = sum(1 for r in recs
                    if (today - timedelta(days=13)).strftime("%Y-%m-%d") <= r["date"] <= (today - timedelta(days=7)).strftime("%Y-%m-%d"))
        result[district] = {"recent": recs[:TOHEO_LIST_ROWS], "daily": daily, "days": days,
                            "last7": last7, "prev7": prev7, "total": len(recs), "ok": ok_any}
    return result


def build_toheo_card(district: str, t: dict) -> str:
    diff = t["last7"] - t["prev7"]
    diff_cls = "ld-up" if diff > 0 else ("ld-down" if diff < 0 else "ld-flat")
    diff_txt = f"{diff:+d}건" if diff else "±0건"
    maxc = max(t["daily"]) if t["daily"] and max(t["daily"]) > 0 else 1
    bars = "".join(
        f'<span class="th-bar" style="height:{max(3, int(26 * c / maxc))}px" title="{d[5:]} · {c}건"></span>'
        for d, c in zip(t["days"], t["daily"]))
    rows_html = ""
    for r in t["recent"]:
        chip = f'<span class="ld-tagchip">{html.escape(r["jimok"])}</span>' if r["jimok"] else ""
        purpose = f'<span class="deal-spec">{html.escape(r["purpose"])}</span>' if r["purpose"] else ""
        rows_html += (f'<div class="deal-row">'
                      f'<span class="deal-date">{r["date"][5:]}</span>'
                      f'<span class="deal-name">{html.escape(r["addr"])}</span>'
                      f'{chip}{purpose}</div>')
    if not rows_html:
        rows_html = ('<div class="deal-row"><span class="deal-empty">'
                     + ("아카이브에 아직 데이터가 없습니다 — 첫 수집 후 누적됩니다."
                        if t["ok"] else "수집 실패 — 아래 링크에서 직접 확인하세요.")
                     + '</span></div>')
    return f"""
        <div class="notion-card deal-card" data-type="land" data-district="{district}">
            <div class="card-meta">
                <span class="tag district-tag">📍 {district}</span>
                <span class="tag toheo-tag">🗂️ 토지거래허가 동향 (수급 지표)</span>
            </div>
            <div class="ld-summary">
                <div class="ld-sum-item">최근 7일 허가 <b>{t["last7"]}건</b> <span class="{diff_cls}">직전 7일 대비 {diff_txt}</span></div>
                <div class="ld-sum-item ld-n">누적 아카이브 {t["total"]}건</div>
            </div>
            <div class="th-bars">{bars}</div>
            <div class="th-bars-label">최근 {TOHEO_TREND_DAYS}일 일별 허가 건수</div>
            <div class="deal-list">{rows_html}</div>
            <div class="card-footer">출처: <a href="{TOHEO_PAGE}" target="_blank">서울부동산정보광장 토지거래허가 내역</a> (K-Geo 연계, 허가일 기준) · 가격정보가 없는 활동량 지표로, 수급 방향성 참고용입니다</div>
        </div>"""


UA_TOHEO = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
            "Referer": TOHEO_PAGE}


def collect_land(today: datetime) -> dict:
    """구별 토지 매매 사례 + 용도지역별 당월/전월 중위단가 요약"""
    result = {d: {"rows": [], "summary": []} for d in DISTRICTS}
    if not MOLIT_KEY:
        print("▶ DATA_GO_KR_KEY 미설정 — 지가분석 수집 생략")
        return result

    months = []
    y, m = today.year, today.month
    for _ in range(LAND_MONTHS):
        months.append(f"{y}{m:02d}")
        m -= 1
        if m == 0:
            y, m = y - 1, 12
    cur_ym, prev_ym = months[0], months[1]

    debug_done = False
    for district in DISTRICTS:
        lawd = LAWD_CD[district]
        print(f"▶ {district} 토지 실거래 수집 중...")
        rows = []
        for ymd in months:
            items = _fetch_deal_xml(LAND_URL, lawd, ymd)
            if items and not debug_done:
                tags = sorted({el.tag for el in items[0]})
                print(f"    [필드 확인] {tags}")
                debug_done = True
            for it in items:
                try:
                    day = datetime(int(_txt(it, "dealYear")), int(_txt(it, "dealMonth")), int(_txt(it, "dealDay")))
                except ValueError:
                    continue
                area = 0.0
                try:
                    area = float(_txt_any(it, ["dealArea", "lndpclAr", "area"]) or 0)
                except ValueError:
                    pass
                amount = _num(_txt(it, "dealAmount"))  # 만원
                if area <= 0 or amount <= 0:
                    continue
                share = _txt_any(it, ["shareDealingType", "dealGbn", "sharDealingType"]).strip()
                is_share = share not in ("", "-", "0", "전체")
                rows.append({
                    "ym": ymd, "date": day,
                    "dong": _txt_any(it, ["umdNm", "dongNm", "sggNm"]),
                    "jimok": _txt_any(it, ["jimok", "lndcgrNm", "landCategory"]),
                    "use": _txt_any(it, ["landUse", "useArea", "useAreaNm", "prposArea1Nm"]),
                    "area": area, "amount": amount,
                    "unit": int(amount * 10000 / area),  # 원/㎡
                    "share": is_share,
                })
        rows.sort(key=lambda x: x["date"], reverse=True)

        # 용도지역별 당월 vs 전월 중위단가 (지분거래 제외, 3건 이상만)
        summary = []
        uses = sorted({r["use"] for r in rows if r["use"]})
        for u in uses:
            cur = [r["unit"] for r in rows if r["use"] == u and r["ym"] == cur_ym and not r["share"]]
            prv = [r["unit"] for r in rows if r["use"] == u and r["ym"] == prev_ym and not r["share"]]
            if len(cur) >= 3 and len(prv) >= 3:
                cm, pm = _median(cur), _median(prv)
                summary.append({"use": u, "cur": int(cm), "chg": (cm - pm) / pm * 100, "n": len(cur)})
            elif len(cur) >= 3:
                summary.append({"use": u, "cur": int(_median(cur)), "chg": None, "n": len(cur)})
        result[district] = {"rows": rows[:MAX_LAND_ROWS], "summary": summary,
                            "total": len(rows), "cur_n": sum(1 for r in rows if r["ym"] == cur_ym)}
        print(f"  → 토지 사례 {len(rows)}건 (당월 {result[district]['cur_n']}건)")
    return result


def build_land_card(district: str, data: dict) -> str:
    rows_html = ""
    for r in data["rows"]:
        d = r["date"]
        share_tag = '<span class="ld-share">지분</span>' if r["share"] else ""
        rows_html += (f'<div class="deal-row">'
                      f'<span class="deal-date">{d.month}/{d.day}</span>'
                      f'<span class="deal-name">{html.escape(r["dong"])}</span>'
                      f'<span class="ld-tagchip">{html.escape(r["jimok"])}</span>'
                      f'<span class="ld-tagchip">{html.escape(r["use"])}</span>{share_tag}'
                      f'<span class="deal-spec">{r["area"]:,.0f}㎡</span>'
                      f'<span class="deal-price">{r["unit"]:,}원/㎡</span>'
                      f'</div>')
    if not rows_html:
        rows_html = '<div class="deal-row"><span class="deal-empty">최근 3개월 토지 매매 사례 없음</span></div>'

    sm_html = ""
    for s in data.get("summary", []):
        if s["chg"] is None:
            chg = '<span class="ld-flat">전월 사례부족</span>'
        else:
            cls = "ld-up" if s["chg"] > 0 else ("ld-down" if s["chg"] < 0 else "ld-flat")
            chg = f'<span class="{cls}">{s["chg"]:+.1f}%</span>'
        sm_html += (f'<div class="ld-sum-item"><b>{html.escape(s["use"])}</b> '
                    f'중위 {s["cur"]:,}원/㎡ {chg} <span class="ld-n">({s["n"]}건)</span></div>')
    if not sm_html:
        sm_html = '<div class="ld-sum-item ld-flat">당월 3건 이상 용도지역 없음 — 중위단가 비교 생략</div>'

    return f"""
        <div class="notion-card deal-card" data-type="land" data-district="{district}">
            <div class="card-meta">
                <span class="tag district-tag">📍 {district}</span>
                <span class="tag land-tag">📐 토지 매매 {data.get("total", 0)}건 (최근 {LAND_MONTHS}개월)</span>
            </div>
            <div class="ld-summary">{sm_html}</div>
            <div class="deal-list">{rows_html}</div>
            <div class="card-footer">출처: 국토교통부 실거래가 · 단가 = 거래금액 ÷ 계약면적 · 중위단가 비교는 지분거래 제외, 당월·전월 각 3건 이상 시 표시 · 지가변동률 조사 참고용 가공자료</div>
        </div>"""




def _load_geocache() -> dict:
    try:
        with open(GEO_CACHE_PATH, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _save_geocache(cache: dict):
    os.makedirs(os.path.dirname(GEO_CACHE_PATH), exist_ok=True)
    with open(GEO_CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False)


def _kakao_geocode(query: str, keyword: bool = False):
    url = ("https://dapi.kakao.com/v2/local/search/keyword.json" if keyword
           else "https://dapi.kakao.com/v2/local/search/address.json")
    try:
        r = requests.get(url, params={"query": query},
                         headers={"Authorization": f"KakaoAK {KAKAO_REST_KEY}"}, timeout=10)
        r.raise_for_status()
        docs = r.json().get("documents", [])
        if docs:
            return float(docs[0]["y"]), float(docs[0]["x"])  # (lat, lng)
    except Exception as e:
        print(f"    [지오코딩 실패] {query}: {type(e).__name__}")
    return None


def apply_geocoding(deals: dict):
    """거래 주소를 좌표로 변환해 각 항목에 lat/lng 부여 (캐시 활용)"""
    if not KAKAO_REST_KEY:
        print("▶ KAKAO_REST_KEY 미설정 — 지도 좌표 변환 생략")
        return
    cache = _load_geocache()
    new_cnt = 0
    for district, items in deals.items():
        for x in items:
            key = f"{district}|{x.get('dong','')}|{x.get('jibun','')}|{x.get('apt','')}"
            if key in cache:
                coord = cache[key]
            else:
                addr = f"서울특별시 {district} {x.get('dong','')} {x.get('jibun','')}".strip()
                coord = _kakao_geocode(addr)
                if coord is None and x.get("apt"):
                    coord = _kakao_geocode(f"{district} {x['apt']}", keyword=True)
                cache[key] = coord
                new_cnt += 1
            if coord:
                x["lat"], x["lng"] = coord
    _save_geocache(cache)
    done = sum(1 for v in deals.values() for x in v if "lat" in x)
    print(f"▶ 지도 좌표 변환: 총 {done}건 표시 가능 (신규 변환 {new_cnt}건, 캐시 재사용)")



def _txt(node, tag):
    el = node.find(tag)
    return el.text.strip() if el is not None and el.text else ""


def _num(s):
    try:
        return int(s.replace(",", "").strip() or 0)
    except ValueError:
        return 0


def _money(man: int) -> str:
    """만원 단위 → '12억 5,000' 형태"""
    eok, rest = divmod(man, 10000)
    if eok and rest:
        return f"{eok}억 {rest:,}"
    if eok:
        return f"{eok}억"
    return f"{rest:,}"


def _fetch_deal_xml(url: str, lawd: str, ymd: str) -> list:
    import xml.etree.ElementTree as ET
    params = {"serviceKey": MOLIT_KEY, "LAWD_CD": lawd, "DEAL_YMD": ymd,
              "numOfRows": "1000", "pageNo": "1"}
    try:
        r = requests.get(url, params=params, timeout=15)
        r.raise_for_status()
        root = ET.fromstring(r.content)
        code = root.findtext(".//resultCode") or ""
        if code not in ("00", "000"):
            print(f"    [API 응답 오류] {root.findtext('.//resultMsg')}")
            return []
        return root.findall(".//item")
    except Exception as e:
        print(f"    [실거래 조회 실패] {type(e).__name__}")
        return []


def _find_progress_file():
    try:
        for f in sorted(os.listdir(PROGRESS_DIR)):
            if f.endswith(".xlsx"):
                return os.path.join(PROGRESS_DIR, f), f
    except FileNotFoundError:
        pass
    return None, None


def load_progress() -> tuple:
    """분기 엑셀 → 구별 정비사업 추진현황. (data dict, 기준 표시 문자열)"""
    path, fname = _find_progress_file()
    result = {d: [] for d in DISTRICTS}
    if not path:
        print("▶ 추진현황 엑셀 없음 (data/*.xlsx) — 탭 생략")
        return result, ""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
    except Exception as e:
        print(f"▶ 추진현황 엑셀 읽기 실패: {type(e).__name__}")
        return result, ""

    # 수동 보정 파일 (선택): {"구역명": {"단계": "...", "메모": "..."}}
    overrides = {}
    try:
        with open(PROGRESS_OVERRIDES, encoding="utf-8") as f:
            overrides = json.load(f)
    except Exception:
        pass

    cnt = 0
    for row in ws.iter_rows(min_row=5, values_only=True):
        gu = row[2]
        if gu not in result:
            continue
        name = str(row[3] or "").strip()
        if not name:
            continue
        stage = str(row[9] or "").strip()
        memo = ""
        if name in overrides:
            stage = overrides[name].get("단계", stage)
            memo = overrides[name].get("메모", "")
        # 현재 단계의 인가/처리 일자
        dcol = STAGE_DATE_COL.get(stage)
        sdate = row[dcol] if dcol is not None and dcol < len(row) else None
        sdate_str = sdate.strftime("%Y-%m-%d") if hasattr(sdate, "strftime") else ""
        hh = row[23]
        result[gu].append({
            "name": name, "loc": str(row[4] or ""),
            "type": str(row[8] or ""), "pub": str(row[6] or ""),
            "stage": stage, "stage_date": sdate_str,
            "households": int(hh) if isinstance(hh, (int, float)) and hh else 0,
            "memo": memo,
        })
        cnt += 1
    # 단계가 늦은(진척된) 구역부터 정렬
    for gu in result:
        result[gu].sort(key=lambda x: (STAGES.index(x["stage"]) if x["stage"] in STAGES else -1, x["name"]), reverse=True)
    asof = fname.replace(".xlsx", "").replace("_", " ").strip()
    print(f"▶ 추진현황 로드: 관할 {cnt}개 구역 (기준: {asof})")
    return result, asof


def build_progress_card(p: dict, district: str) -> str:
    idx = STAGES.index(p["stage"]) if p["stage"] in STAGES else -1
    segs = "".join(
        f'<span class="pg-seg{" pg-on" if i <= idx else ""}" title="{s}"></span>'
        for i, s in enumerate(STAGES))
    hh = f' · 건립 {p["households"]:,}세대' if p["households"] else ""
    memo = f'<div class="pg-memo">📝 {html.escape(p["memo"])}</div>' if p["memo"] else ""
    date_part = f' ({p["stage_date"]})' if p["stage_date"] else ""
    return f"""
        <div class="notion-card" data-type="prog" data-district="{district}">
            <div class="card-meta">
                <span class="tag district-tag">📍 {district}</span>
                <span class="tag prog-tag">🏗️ {html.escape(p['type'])}</span>
                <span class="tag date-tag">{html.escape(p['pub'])}</span>
            </div>
            <h3 class="news-title">{html.escape(p['name'])}</h3>
            <div class="pg-bar">{segs}</div>
            <div class="pg-label">현재 단계: <b>{html.escape(p['stage'])}</b>{date_part}{hh}</div>
            {memo}
            <div class="card-footer">{html.escape(p['loc'])}</div>
        </div>"""


def collect_deals(today: datetime) -> dict:
    cutoff = today - timedelta(days=DEAL_DAYS_BACK)
    result = {d: [] for d in DISTRICTS}
    if not MOLIT_KEY:
        print("▶ DATA_GO_KR_KEY 미설정 — 실거래 수집 생략")
        return result

    # 이번 달 + (기간이 지난달에 걸치면) 지난달 조회
    months = {today.strftime("%Y%m")}
    months.add(cutoff.strftime("%Y%m"))

    for district in DISTRICTS:
        lawd = LAWD_CD[district]
        print(f"▶ {district} 실거래 수집 중...")
        deals = []
        for ymd in sorted(months):
            # 매매
            for it in _fetch_deal_xml(TRADE_URL, lawd, ymd):
                try:
                    day = datetime(int(_txt(it, "dealYear")), int(_txt(it, "dealMonth")), int(_txt(it, "dealDay")))
                except ValueError:
                    continue
                if not (cutoff <= day <= today):
                    continue
                deals.append({
                    "type": "매매", "date": day,
                    "apt": _txt(it, "aptNm") or _txt(it, "offiNm"),
                    "dong": _txt(it, "umdNm"),
                    "jibun": _txt(it, "jibun"),
                    "area": _txt(it, "excluUseAr"),
                    "floor": _txt(it, "floor"),
                    "price": _money(_num(_txt(it, "dealAmount"))),
                })
            # 전월세
            for it in _fetch_deal_xml(RENT_URL, lawd, ymd):
                try:
                    day = datetime(int(_txt(it, "dealYear")), int(_txt(it, "dealMonth")), int(_txt(it, "dealDay")))
                except ValueError:
                    continue
                if not (cutoff <= day <= today):
                    continue
                monthly = _num(_txt(it, "monthlyRent"))
                deposit = _num(_txt(it, "deposit"))
                deals.append({
                    "type": "월세" if monthly else "전세", "date": day,
                    "apt": _txt(it, "aptNm"),
                    "dong": _txt(it, "umdNm"),
                    "jibun": _txt(it, "jibun"),
                    "area": _txt(it, "excluUseAr"),
                    "floor": _txt(it, "floor"),
                    "price": (f"{_money(deposit)}/{monthly:,}" if monthly else _money(deposit)),
                })
        deals.sort(key=lambda x: x["date"], reverse=True)
        result[district] = deals[:MAX_DEALS_PER_DISTRICT]
        print(f"  → 실거래 {len(result[district])}건 (매매/전세/월세)")
    return result


def build_deal_card(district: str, deals: list, today: datetime) -> str:
    if not deals:
        rows = '<div class="deal-row"><span class="deal-empty">최근 7일 계약분 없음 (신고 지연분은 이후 반영될 수 있음)</span></div>'
    else:
        rows = ""
        for x in deals:
            d = x["date"]
            area = x["area"].split(".")[0] + "㎡" if x["area"] else ""
            rows += (f'<div class="deal-row">'
                     f'<span class="deal-type deal-{x["type"]}">{x["type"]}</span>'
                     f'<span class="deal-name">{html.escape(x["dong"])} {html.escape(x["apt"])}</span>'
                     f'<span class="deal-spec">{area} {x["floor"]}층</span>'
                     f'<span class="deal-price">{x["price"]}</span>'
                     f'<span class="deal-date">{d.month}/{d.day}</span>'
                     f'</div>')
    return f"""
        <div class="notion-card deal-card" data-type="deal" data-district="{district}">
            <div class="card-meta">
                <span class="tag district-tag">📍 {district}</span>
                <span class="tag deal-tag">🏠 아파트 실거래 {len(deals)}건</span>
            </div>
            <div class="deal-list">{rows}</div>
            <div class="card-footer">출처: 국토교통부 실거래가 공개시스템 · 계약일 기준 최근 {DEAL_DAYS_BACK}일 (금액 단위: 만원)</div>
        </div>"""


# ──────────────────────────────────────────────
# 뉴스 수집 (네이버 검색 API) — v2와 동일
# ──────────────────────────────────────────────
API_URL = "https://openapi.naver.com/v1/search/news.json"
NAVER_HEADERS = {
    "X-Naver-Client-Id": NAVER_CLIENT_ID,
    "X-Naver-Client-Secret": NAVER_CLIENT_SECRET,
}
TAG_RE = re.compile(r"</?b>")


def clean(text: str) -> str:
    return html.unescape(TAG_RE.sub("", text)).strip()


def fetch_news(district: str, keyword: str) -> list:
    params = {"query": f"{district} {keyword}", "display": MAX_PER_QUERY, "sort": "date"}
    try:
        r = requests.get(API_URL, headers=NAVER_HEADERS, params=params, timeout=10)
        r.raise_for_status()
        return r.json().get("items", [])
    except Exception as e:
        print(f"  [경고] '{district} {keyword}' 검색 실패: {e}")
        return []


def collect_news(today: datetime) -> dict:
    cutoff = today - timedelta(days=DAYS_BACK)
    seen_links = set()
    result = {d: [] for d in DISTRICTS}

    for district in DISTRICTS:
        print(f"▶ {district} 뉴스 수집 중...")
        articles = []
        for kw in KEYWORDS:
            for item in fetch_news(district, kw):
                link = item.get("originallink") or item.get("link", "")
                naver_link = item.get("link", "")
                if not link or link in seen_links or naver_link in seen_links:
                    continue
                try:
                    pub = parsedate_to_datetime(item["pubDate"]).astimezone(KST).replace(tzinfo=None)
                except Exception:
                    continue
                if pub < cutoff:
                    continue
                title = clean(item.get("title", ""))
                desc = clean(item.get("description", ""))
                if district not in title + desc:
                    continue
                if not any(w in title + desc for w in RELEVANCE_WORDS):
                    continue
                seen_links.add(link)
                seen_links.add(naver_link)
                articles.append({"district": district, "title": title,
                                 "link": naver_link or link, "summary": desc, "date": pub})
        articles.sort(key=lambda a: a["date"], reverse=True)
        result[district] = articles[:MAX_PER_DISTRICT]
        print(f"  → 뉴스 {len(result[district])}건 채택")
    return result


# ──────────────────────────────────────────────
# HTML 생성 (뉴스/고시공고 탭)
# ──────────────────────────────────────────────
def make_summary_bullets(summary: str) -> str:
    parts = re.split(r"(?<=[.다요음됨함])\s+", summary)
    parts = [p.strip().rstrip(".") for p in parts if len(p.strip()) > 10][:2]
    if not parts:
        parts = [summary[:120]]
    return "<br>".join(f"• {html.escape(p)}" for p in parts)


def build_news_card(a: dict) -> str:
    d = a["date"]
    return f"""
        <div class="notion-card" data-type="news" data-district="{a['district']}">
            <div class="card-meta">
                <span class="tag district-tag">📍 {a['district']}</span>
                <span class="tag source-tag">📰 네이버뉴스</span>
                <span class="tag date-tag">📅 {d.year}년 {d.month}월 {d.day}일</span>
            </div>
            <h3 class="news-title"><a href="{a['link']}" target="_blank">{html.escape(a['title'])}</a></h3>
            <div class="summary-box">{make_summary_bullets(a['summary'])}</div>
            <div class="card-footer"><a href="{a['link']}" target="_blank">상세 원문 보기 →</a></div>
        </div>"""


def build_notice_card(district: str) -> str:
    b = NOTICE_BOARDS[district]
    return f"""
        <div class="notion-card" data-type="notice" data-district="{district}">
            <div class="card-meta">
                <span class="tag district-tag">📍 {district}</span>
                <span class="tag notice-tag">📢 고시공고</span>
            </div>
            <h3 class="news-title"><a href="{b['url']}" target="_blank">{district}청 {b['name']}</a></h3>
            <div class="summary-box">• 구청 공식 홈페이지 고시공고 게시판으로 이동합니다.<br>• 정비사업 관련 고시문 원문과 첨부파일(PDF/HWP)을 게시판에서 확인하세요.</div>
            <div class="card-footer"><a href="{b['url']}" target="_blank">게시판 바로가기 →</a></div>
        </div>"""


def build_html(news: dict, deals: dict, progress: dict, prog_asof: str, land: dict, toheo: dict, today: datetime) -> str:
    counts = {"news": {"all": sum(len(v) for v in news.values()),
                       **{d: len(news[d]) for d in DISTRICTS}},
              "deal": {"all": sum(len(v) for v in deals.values()),
                       **{d: len(deals[d]) for d in DISTRICTS}},
              "prog": {"all": sum(len(v) for v in progress.values()),
                       **{d: len(progress[d]) for d in DISTRICTS}},
              "land": {"all": sum(land[d].get("total", 0) for d in DISTRICTS),
                       **{d: land[d].get("total", 0) for d in DISTRICTS}}}

    all_news = sorted((a for v in news.values() for a in v), key=lambda x: x["date"], reverse=True)
    cards = "".join(build_news_card(a) for a in all_news) + \
            "".join(build_notice_card(d) for d in DISTRICTS) + \
            "".join(build_deal_card(d, deals[d], today) for d in DISTRICTS) + \
            "".join(build_progress_card(p, d) for d in DISTRICTS for p in progress[d]) + \
            "".join(build_land_card(d, land[d]) for d in DISTRICTS) + \
            "".join(build_toheo_card(d, toheo[d]) for d in DISTRICTS)

    sidebar = ['<div class="sidebar-item active" data-district="all">🌐 전체</div>']
    sidebar += [f'<div class="sidebar-item" data-district="{d}">📍 {d}</div>' for d in DISTRICTS]

    deal_points = [
        {"d": dist, "t": x["type"], "n": f"{x['dong']} {x['apt']}",
         "p": x["price"], "dt": f"{x['date'].month}/{x['date'].day}",
         "lat": x["lat"], "lng": x["lng"]}
        for dist, v in deals.items() for x in v if "lat" in x
    ]
    deals_json = json.dumps(deal_points, ensure_ascii=False)

    _n = html.escape(UPDATE_NOTICE.strip())
    _dur = max(12, int(len(UPDATE_NOTICE) * 0.45))  # 문구가 길수록 천천히
    notice_bar = ((f'<div class="update-bar"><div class="update-track" style="animation-duration:{_dur}s">'
                   f'<span class="update-item">📌 {_n}</span><span class="update-item">📌 {_n}</span>'
                   f'</div></div>')
                  if UPDATE_NOTICE.strip() else "")
    date_str = today.strftime("%Y-%m-%d")
    period_str = (today - timedelta(days=DAYS_BACK)).strftime("%Y-%m-%d")
    updated_str = today.strftime("%Y-%m-%d %H:%M")

    page = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <meta name="theme-color" content="#fbfbfa">
    <meta name="apple-mobile-web-app-capable" content="yes">
    <meta name="apple-mobile-web-app-title" content="동부 AI toolkit">
    <link rel="manifest" href="manifest.json">
    <title>서울동부지사 AI toolkit</title>
    <style>
        body {{ margin: 0; padding: 0; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Pretendard, "Malgun Gothic", sans-serif; background-color: #fbfbfa; color: #37352f; }}
        #header {{ padding: 40px 50px 0 50px; background-color: #fbfbfa; border-bottom: 1px solid #edf2fa; }}
        #header h1 {{ font-size: 28px; font-weight: 700; margin: 0 0 8px 0; }}
        #header .subtitle {{ color: #73726e; font-size: 14px; margin-bottom: 12px; }}
        .update-bar {{ background-color: #fdf6e3; border: 1px solid #f0e2b6; color: #6b5a1e; font-size: 13px; padding: 7px 0; border-radius: 6px; margin-bottom: 14px; overflow: hidden; }}
        .update-track {{ display: inline-flex; white-space: nowrap; animation: marquee linear infinite; will-change: transform; }}
        .update-item {{ padding-right: 80px; }}
        .update-bar:hover .update-track {{ animation-play-state: paused; }}
        @keyframes marquee {{ from {{ transform: translateX(0); }} to {{ transform: translateX(-50%); }} }}
        .tab-bar {{ display: flex; flex-direction: column; }}
        .tab-row {{ display: flex; gap: 4px; }}
        .tab-btn {{ padding: 10px 18px; font-size: 15px; font-weight: 600; color: #73726e; cursor: pointer; border-bottom: 2px solid transparent; }}
        .tab-btn.active {{ color: #37352f; border-bottom-color: #37352f; }}
        #container {{ display: flex; padding: 0 50px 50px 50px; }}
        #sidebar {{ width: 200px; padding-right: 25px; border-right: 1px solid #eaeaea; margin-top: 30px; flex-shrink: 0; }}
        .sidebar-title {{ font-size: 11px; font-weight: 600; color: #acaba9; margin-bottom: 12px; padding-left: 10px; letter-spacing: 0.5px; }}
        .sidebar-item {{ padding: 8px 12px; font-size: 14px; border-radius: 6px; margin-bottom: 4px; cursor: pointer; transition: background 0.2s; }}
        .sidebar-item:hover {{ background-color: #f1f1ef; }}
        .sidebar-item.active {{ background-color: #ececed; font-weight: 600; }}
        #main-content {{ flex: 1; padding-left: 40px; margin-top: 30px; min-width: 0; }}
        .view-bar {{ font-size: 14px; font-weight: 500; margin-bottom: 20px; border-bottom: 1px solid #eaeaea; padding-bottom: 8px; }}
        .card-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(360px, 1fr)); gap: 20px; }}
        .notion-card {{ background: #fff; border: 1px solid #e9e9e6; border-radius: 8px; padding: 20px; box-shadow: 0 1px 3px rgba(0,0,0,0.04); display: flex; flex-direction: column; justify-content: space-between; transition: transform 0.2s, box-shadow 0.2s; }}
        .notion-card:hover {{ transform: translateY(-2px); box-shadow: 0 4px 12px rgba(0,0,0,0.06); }}
        .fallback-card {{ border-style: dashed; background: #fffdf5; }}
        .card-meta {{ display: flex; gap: 8px; margin-bottom: 12px; align-items: center; flex-wrap: wrap; }}
        .tag {{ font-size: 12px; font-weight: 500; padding: 3px 8px; border-radius: 4px; white-space: nowrap; }}
        .district-tag {{ background-color: #edf2fa; color: #1f497d; }}
        .source-tag {{ background-color: #fbe4e4; color: #9c3a3a; }}
        .notice-tag {{ background-color: #fdecc8; color: #8a6116; }}
        .dept-tag {{ background-color: #e8f3ec; color: #2b6a46; }}
        .date-tag {{ background-color: #f1f1ef; color: #5a5a57; }}
        .news-title {{ font-size: 16px; font-weight: 600; margin: 0 0 14px 0; line-height: 1.4; }}
        .news-title a {{ color: #2383e2; text-decoration: none; }}
        .news-title a:hover {{ text-decoration: underline; }}
        .summary-box {{ background-color: #f7f7f5; padding: 14px; border-radius: 6px; font-size: 13.5px; line-height: 1.6; margin-bottom: 14px; flex-grow: 1; }}
        .card-footer {{ font-size: 13px; margin-top: 8px; border-top: 1px dashed #eaeaea; padding-top: 10px; }}
        .card-footer a {{ color: #2383e2; text-decoration: none; font-weight: 500; }}
        .card-footer a:hover {{ text-decoration: underline; }}
        .deal-card {{ grid-column: 1 / -1; }}
        .deal-list {{ background-color: #f7f7f5; border-radius: 6px; padding: 6px 10px; margin-bottom: 12px; }}
        .deal-row {{ display: flex; align-items: center; gap: 10px; padding: 7px 4px; border-bottom: 1px solid #ececea; font-size: 13.5px; flex-wrap: wrap; }}
        .deal-row:last-child {{ border-bottom: none; }}
        .deal-type {{ font-size: 11.5px; font-weight: 700; padding: 2px 7px; border-radius: 4px; flex-shrink: 0; }}
        .deal-매매 {{ background-color: #fbe4e4; color: #9c3a3a; }}
        .deal-전세 {{ background-color: #e3ecf7; color: #1f497d; }}
        .deal-월세 {{ background-color: #e8f3ec; color: #2b6a46; }}
        .deal-name {{ font-weight: 600; min-width: 0; }}
        .deal-spec {{ color: #73726e; flex-shrink: 0; }}
        .deal-price {{ font-weight: 700; margin-left: auto; flex-shrink: 0; }}
        .deal-date {{ color: #acaba9; font-size: 12px; width: 34px; text-align: right; flex-shrink: 0; }}
        .deal-empty {{ color: #73726e; font-size: 13px; }}
        .deal-tag {{ background-color: #f3e8f7; color: #6b3f85; }}
        #deal-map-wrap {{ display: none; margin-bottom: 20px; }}
        #deal-map {{ width: 100%; height: 360px; border-radius: 8px; border: 1px solid #e9e9e6; }}
        .map-legend {{ font-size: 12.5px; color: #73726e; margin-top: 6px; }}
        .lg-매매 {{ color: #d94343; }} .lg-전세 {{ color: #2f6bd8; }} .lg-월세 {{ color: #2b8a4e; }}
        .prog-tag {{ background-color: #e3ecf7; color: #1f497d; }}
        .pg-bar {{ display: flex; gap: 3px; margin: 4px 0 10px 0; }}
        .pg-seg {{ flex: 1; height: 8px; border-radius: 4px; background-color: #ececea; }}
        .pg-on {{ background-color: #2f6bd8; }}
        .pg-label {{ font-size: 13px; color: #37352f; margin-bottom: 10px; }}
        .pg-memo {{ font-size: 12.5px; color: #8a6116; background-color: #fdf6e3; border-radius: 4px; padding: 5px 8px; margin-bottom: 10px; }}
        .land-tag {{ background-color: #e8ecf3; color: #3a4a6b; }}
        .ld-summary {{ display: flex; flex-wrap: wrap; gap: 8px 18px; background-color: #f1f4f9; border-radius: 6px; padding: 10px 12px; margin-bottom: 12px; }}
        .ld-sum-item {{ font-size: 13px; }}
        .ld-up {{ color: #d94343; font-weight: 700; }}
        .ld-down {{ color: #2f6bd8; font-weight: 700; }}
        .ld-flat {{ color: #73726e; }}
        .ld-n {{ color: #acaba9; font-size: 12px; }}
        .ld-tagchip {{ font-size: 11.5px; background-color: #f1f1ef; color: #5a5a57; padding: 2px 6px; border-radius: 4px; flex-shrink: 0; }}
        .ld-share {{ font-size: 11.5px; background-color: #fdecc8; color: #8a6116; padding: 2px 6px; border-radius: 4px; flex-shrink: 0; }}
        .deal-date {{ color: #acaba9; font-size: 12px; width: 34px; flex-shrink: 0; }}
        #lab-box {{ display: none; text-align: center; padding: 90px 20px; }}
        .lab-icon {{ font-size: 64px; margin-bottom: 18px; }}
        .lab-text {{ font-size: 15px; color: #73726e; }}
        #lab-box {{ text-align: left; padding: 10px 0 40px 0; max-width: 760px; }}
        .rt-head {{ font-size: 17px; margin-bottom: 8px; }}
        .rt-beta {{ font-size: 11px; background-color: #f3e8f7; color: #6b3f85; padding: 2px 7px; border-radius: 4px; vertical-align: middle; }}
        .rt-guide {{ font-size: 13px; color: #73726e; margin-bottom: 10px; line-height: 1.6; }}
        #rt-input {{ width: 100%; box-sizing: border-box; padding: 12px; border: 1px solid #e0e0dd; border-radius: 8px; font-size: 13.5px; font-family: inherit; resize: vertical; }}
        .rt-controls {{ display: flex; justify-content: space-between; align-items: center; margin: 10px 0; font-size: 13px; }}
        #rt-btn {{ padding: 9px 18px; border: none; border-radius: 8px; background-color: #37352f; color: #fff; font-weight: 600; cursor: pointer; }}
        #rt-btn:hover {{ background-color: #1f1e1b; }}
        #rt-msg {{ min-height: 20px; font-size: 13px; color: #37352f; margin-bottom: 8px; }}
        #rt-map {{ display: none; width: 100%; height: 360px; border-radius: 8px; border: 1px solid #e9e9e6; margin-bottom: 12px; }}
        #rt-list {{ font-size: 13.5px; line-height: 1.9; padding-left: 22px; margin: 0 0 8px 0; }}
        .rt-km {{ color: #2f6bd8; font-size: 12px; font-weight: 600; }}
        .rt-note {{ font-size: 12px; color: #acaba9; }}
        .rt-pin {{ background-color: #2f6bd8; color: #fff; font-size: 11.5px; font-weight: 700; border-radius: 999px; padding: 3px 8px; border: 2px solid #fff; box-shadow: 0 1px 4px rgba(0,0,0,0.3); }}
        .rt-pin-start {{ background-color: #d94343; }}
        .toheo-tag {{ background-color: #fdecc8; color: #8a6116; }}
        .th-bars {{ display: flex; align-items: flex-end; gap: 3px; height: 30px; margin-bottom: 2px; }}
        .th-bar {{ width: 9px; background-color: #8aa8d8; border-radius: 2px 2px 0 0; }}
        .th-bars-label {{ font-size: 11.5px; color: #acaba9; margin-bottom: 10px; }}

        @media (max-width: 768px) {{
            #header {{ padding: 24px 16px 0 16px; }}
            #header h1 {{ font-size: 20px; }}
            #header .subtitle {{ font-size: 12px; }}
            .update-bar {{ font-size: 12px; padding: 6px 10px; }}
            .tab-btn {{ flex: 1; text-align: center; padding: 10px 4px; font-size: 14px; }}
            #container {{ flex-direction: column; padding: 0 16px 30px 16px; }}
            #sidebar {{ width: 100%; padding-right: 0; border-right: none; border-bottom: 1px solid #eaeaea; margin-top: 12px; padding-bottom: 10px; position: sticky; top: 0; background-color: #fbfbfa; z-index: 10; }}
            .sidebar-title {{ display: none; }}
            #sidebar-items {{ display: flex; overflow-x: auto; gap: 6px; -webkit-overflow-scrolling: touch; scrollbar-width: none; }}
            #sidebar-items::-webkit-scrollbar {{ display: none; }}
            .sidebar-item {{ white-space: nowrap; margin-bottom: 0; background-color: #f1f1ef; font-size: 13px; padding: 7px 12px; border-radius: 16px; }}
            .sidebar-item.active {{ background-color: #37352f; color: #fff; }}
            #main-content {{ padding-left: 0; margin-top: 16px; }}
            .card-grid {{ grid-template-columns: 1fr; gap: 14px; }}
            #deal-map {{ height: 260px; }}
            .notion-card {{ padding: 16px; }}
            .notion-card:hover {{ transform: none; }}
        }}
    </style>
</head>
<body>
__GATE__
    <div id="header">
        <h1>서울동부지사 AI toolkit</h1>
        <div class="subtitle">📅 {date_str} 기준 · 최근 {DAYS_BACK}일 ({period_str} ~) · 갱신 {updated_str} KST</div>
        {notice_bar}
        <div class="tab-bar">
            <div class="tab-row">
                <div class="tab-btn active" data-tab="news">📰 뉴스</div>
                <div class="tab-btn" data-tab="notice">📢 고시공고</div>
                <div class="tab-btn" data-tab="deal">🏠 최근 실거래</div>
            </div>
            <div class="tab-row">
                <div class="tab-btn" data-tab="prog">🏗️ 추진현황</div>
                <div class="tab-btn" data-tab="land">📐 지가분석</div>
                <div class="tab-btn" data-tab="lab">🧪 실험실</div>
            </div>
        </div>
    </div>

    <div id="container">
        <div id="sidebar">
            <div class="sidebar-title">관할 자치구</div>
            <div id="sidebar-items">{"".join(sidebar)}</div>
        </div>
        <div id="main-content">
            <div class="view-bar" id="view-bar"></div>
            <div id="deal-map-wrap">
                <div id="deal-map"></div>
                <div class="map-legend"><span class="lg lg-매매">● 매매</span> <span class="lg lg-전세">● 전세</span> <span class="lg lg-월세">● 월세</span> — 핀을 누르면 상세 표시</div>
            </div>
            <div id="lab-box">__LAB_HTML__</div>
            <div class="card-grid">{cards}</div>
        </div>
    </div>

    <script>
        const COUNTS = {json.dumps(counts, ensure_ascii=False)};
        let tab = 'news', district = 'all';

        function render() {{
            document.querySelectorAll('.notion-card').forEach(c => {{
                const show = c.dataset.type === tab && (district === 'all' || c.dataset.district === district);
                c.style.display = show ? 'flex' : 'none';
            }});
            document.querySelectorAll('.tab-btn').forEach(b =>
                b.classList.toggle('active', b.dataset.tab === tab));
            document.querySelectorAll('.sidebar-item').forEach(s => {{
                s.classList.toggle('active', s.dataset.district === district);
                const name = s.dataset.district === 'all' ? '🌐 전체' : '📍 ' + s.dataset.district;
                s.textContent = (tab === 'notice' || tab === 'lab') ? name : name + ' (' + (COUNTS[tab][s.dataset.district] || 0) + ')';
            }});
            updateDealMap();
            document.getElementById('lab-box').style.display = tab === 'lab' ? 'block' : 'none';
            document.getElementById('view-bar').textContent =
                tab === 'news' ? '📋 뉴스 갤러리 — 최신순' : tab === 'notice' ? '📋 구청별 고시공고 게시판 바로가기' : tab === 'deal' ? '📋 구별 아파트 실거래 — 계약일 기준 최근 7일' : tab === 'prog' ? '📋 정비사업 추진현황 — __PROG_ASOF__ · 진척 단계순' : tab === 'land' ? '📋 토지 매매 사례 분석 — 지가변동률 조사 지원 (최신순)' : '🧪 실험실 — 준비 중인 기능';
        }}

        document.querySelectorAll('.tab-btn').forEach(b =>
            b.addEventListener('click', () => {{ tab = b.dataset.tab; render(); }}));
        document.querySelectorAll('.sidebar-item').forEach(s =>
            s.addEventListener('click', () => {{ district = s.dataset.district; render(); }}));
        __DEAL_MAP_JS__
        render();
    </script>
</body>
</html>"""
    hashes = _auth_hashes()
    logo = ('<div class="gate-logo gate-logo-google" role="img" aria-label="symbol"></div>'
            if os.path.exists(os.path.join("docs", "logo.png")) else GATE_LOGO_FALLBACK)
    gate = (GATE_BLOCK.replace("__AUTH_HASHES__", json.dumps(hashes)).replace("__GATE_LOGO__", logo)
            if hashes else "")
    page = page.replace("__GATE__", gate)
    page = page.replace("__PROG_ASOF__", html.escape(prog_asof) or "기준 파일 없음")
    page = page.replace("__LAB_HTML__", LAB_ROUTE_HTML if KAKAO_JS_KEY else LAB_PLACEHOLDER)
    page = page.replace("__DEAL_MAP_JS__", (DEAL_MAP_JS + LAB_ROUTE_JS) if KAKAO_JS_KEY else "function updateDealMap(){}")
    page = page.replace("__DEALS__", deals_json).replace("__KAKAO_JS_KEY__", KAKAO_JS_KEY)
    return page


LAB_PLACEHOLDER = """
                <div class="lab-icon">🔬</div>
                <div class="lab-text">아직 실험중이에요.</div>"""

LAB_ROUTE_HTML = """
                <div class="rt-head">🧭 <b>출장 표본지 최적동선</b> <span class="rt-beta">실험실 β</span></div>
                <div class="rt-guide">표본지 주소를 한 줄에 하나씩 입력하세요. <b>첫 줄이 출발지</b>로 고정되고, 나머지 방문 순서를 최적화합니다. (지번·도로명 모두 가능, 예: 노원구 상계동 666-12)</div>
                <textarea id="rt-input" rows="7" placeholder="서울 노원구 노해로 437 (출발지)&#10;노원구 상계동 666-12&#10;도봉구 창동 135-8&#10;강북구 미아동 234-5"></textarea>
                <div class="rt-controls">
                    <label><input type="checkbox" id="rt-round"> 출발지로 복귀(왕복)</label>
                    <button id="rt-btn">최적 동선 계산</button>
                </div>
                <div id="rt-msg"></div>
                <div id="rt-map"></div>
                <ol id="rt-list"></ol>
                <div class="rt-note">※ 거리는 직선거리 기준 근사치입니다. 실제 도로·교통 상황에 따른 소요시간은 내비게이션으로 확인하세요.</div>"""

GATE_LOGO_FALLBACK = """<svg class="gate-logo" viewBox="0 0 100 100" xmlns="http://www.w3.org/2000/svg">
            <path d="M50 8 A42 42 0 0 1 92 50" fill="none" stroke="#4285F4" stroke-width="13" stroke-linecap="round"/>
            <path d="M92 50 A42 42 0 0 1 50 92" fill="none" stroke="#34A853" stroke-width="13" stroke-linecap="round"/>
            <path d="M50 92 A42 42 0 0 1 8 50" fill="none" stroke="#FBBC05" stroke-width="13" stroke-linecap="round"/>
            <path d="M8 50 A42 42 0 0 1 50 8" fill="none" stroke="#EA4335" stroke-width="13" stroke-linecap="round"/>
        </svg>"""

GATE_BLOCK = r"""
    <div id="gate">
__GATE_LOGO__
        <div class="gate-box">
            <input id="gate-id" type="text" placeholder="아이디" autocomplete="username">
            <input id="gate-pw" type="password" placeholder="비밀번호" autocomplete="current-password">
            <button id="gate-btn">접속</button>
            <div id="gate-msg"></div>
        </div>
    </div>
    <style>
        #gate { position: fixed; inset: 0; z-index: 9999; background-color: #fbfbfa;
                display: flex; flex-direction: column; align-items: center; justify-content: center; gap: 34px; }
        .gate-logo { width: 92px; height: 92px; object-fit: contain; }
        .gate-logo-google {
            background: conic-gradient(from -45deg, #4285F4 0 25%, #EA4335 25% 50%, #FBBC05 50% 75%, #34A853 75% 100%);
            -webkit-mask: url(logo.png) center / contain no-repeat;
            mask: url(logo.png) center / contain no-repeat;
        }
        .gate-box { display: flex; flex-direction: column; gap: 10px; width: 240px; }
        .gate-box input { padding: 11px 14px; border: 1px solid #e0e0dd; border-radius: 8px; font-size: 14px;
                          background-color: #ffffff; outline: none; }
        .gate-box input:focus { border-color: #4285F4; }
        #gate-btn { padding: 11px; border: none; border-radius: 8px; background-color: #37352f; color: #fff;
                    font-size: 14px; font-weight: 600; cursor: pointer; }
        #gate-btn:hover { background-color: #1f1e1b; }
        #gate-msg { min-height: 18px; font-size: 12.5px; color: #d94343; text-align: center; }
    </style>
    <script>
    (function(){
        const HASHES = __AUTH_HASHES__;
        const KEY = 'toolkit_auth';
        const gate = document.getElementById('gate');
        async function sha(t){
            const b = await crypto.subtle.digest('SHA-256', new TextEncoder().encode(t));
            return [...new Uint8Array(b)].map(x => x.toString(16).padStart(2,'0')).join('');
        }
        function pass(h){
            try { localStorage.setItem(KEY, JSON.stringify({h: h, exp: Date.now() + 7*24*3600*1000})); } catch(e) {}
            gate.remove();
        }
        try {
            const s = JSON.parse(localStorage.getItem(KEY) || 'null');
            if (s && s.exp > Date.now() && HASHES.includes(s.h)) { gate.remove(); return; }
        } catch(e) {}
        async function tryLogin(){
            const id = document.getElementById('gate-id').value.trim();
            const pw = document.getElementById('gate-pw').value;
            const h = await sha(id + ':' + pw);
            if (HASHES.includes(h)) { pass(h); }
            else {
                document.getElementById('gate-msg').textContent = '아이디 또는 비밀번호가 올바르지 않습니다.';
                document.getElementById('gate-pw').value = '';
            }
        }
        document.getElementById('gate-btn').addEventListener('click', tryLogin);
        document.getElementById('gate-pw').addEventListener('keydown', e => { if (e.key === 'Enter') tryLogin(); });
    })();
    </script>
"""


DEAL_MAP_JS = r"""
        const DEALS = __DEALS__;
        const PIN_COLOR = {"매매": "#d94343", "전세": "#2f6bd8", "월세": "#2b8a4e"};
        let map = null, markers = [], infoWin = null, sdkLoaded = false;

        function pinImage(t) {
            const svg = '<svg xmlns="http://www.w3.org/2000/svg" width="18" height="18">' +
                '<circle cx="9" cy="9" r="7" fill="' + PIN_COLOR[t] + '" stroke="white" stroke-width="2.5"/></svg>';
            return new kakao.maps.MarkerImage('data:image/svg+xml;utf8,' + encodeURIComponent(svg),
                new kakao.maps.Size(18, 18), {offset: new kakao.maps.Point(9, 9)});
        }

        function initMap() {
            map = new kakao.maps.Map(document.getElementById('deal-map'),
                {center: new kakao.maps.LatLng(37.60, 127.06), level: 7});
            infoWin = new kakao.maps.InfoWindow({removable: true});
            DEALS.forEach(x => {
                const m = new kakao.maps.Marker({
                    position: new kakao.maps.LatLng(x.lat, x.lng), image: pinImage(x.t)});
                kakao.maps.event.addListener(m, 'click', () => {
                    infoWin.setContent('<div style="padding:8px 12px;font-size:12.5px;line-height:1.5;">' +
                        '<b>' + x.n + '</b><br>' + x.t + ' ' + x.p + ' · ' + x.dt + ' 계약</div>');
                    infoWin.open(map, m);
                });
                m._d = x.d; markers.push(m);
            });
        }

        function updateDealMap() {
            const wrap = document.getElementById('deal-map-wrap');
            wrap.style.display = (tab === 'deal' && DEALS.length) ? 'block' : 'none';
            if (tab !== 'deal' || !DEALS.length) return;
            if (!sdkLoaded) return;  // SDK 로드 후 재호출됨
            if (!map) initMap();
            map.relayout();
            const bounds = new kakao.maps.LatLngBounds();
            let visible = 0;
            markers.forEach(m => {
                const show = district === 'all' || m._d === district;
                m.setMap(show ? map : null);
                if (show) { bounds.extend(m.getPosition()); visible++; }
            });
            if (infoWin) infoWin.close();
            if (visible) map.setBounds(bounds, 30);
        }

        (function loadKakao() {
            if (!DEALS.length) return;
            const s = document.createElement('script');
            s.src = 'https://dapi.kakao.com/v2/maps/sdk.js?appkey=__KAKAO_JS_KEY__&autoload=false&libraries=services';
            s.onload = () => kakao.maps.load(() => { sdkLoaded = true; updateDealMap(); });
            document.head.appendChild(s);
        })();
"""


LAB_ROUTE_JS = r"""
        function haver(a, b) {
            const R = 6371, dLa = (b.lat - a.lat) * Math.PI / 180, dLo = (b.lng - a.lng) * Math.PI / 180;
            const h = Math.sin(dLa/2)**2 + Math.cos(a.lat*Math.PI/180) * Math.cos(b.lat*Math.PI/180) * Math.sin(dLo/2)**2;
            return 2 * R * Math.asin(Math.sqrt(h));
        }
        function optimizeRoute(pts, roundTrip) {
            const n = pts.length, D = pts.map(p => pts.map(q => haver(p, q)));
            let order = [0]; const left = new Set([...Array(n).keys()].slice(1));
            while (left.size) {  // 최근접 이웃
                const last = order[order.length - 1];
                let best = null, bd = Infinity;
                left.forEach(i => { if (D[last][i] < bd) { bd = D[last][i]; best = i; } });
                order.push(best); left.delete(best);
            }
            const cost = o => {  // 2-opt 개선
                let s = 0;
                for (let i = 0; i < o.length - 1; i++) s += D[o[i]][o[i+1]];
                if (roundTrip) s += D[o[o.length-1]][o[0]];
                return s;
            };
            let improved = true;
            while (improved) {
                improved = false;
                for (let i = 1; i < order.length - 1; i++)
                    for (let k = i + 1; k < order.length; k++) {
                        const cand = order.slice(0, i).concat(order.slice(i, k+1).reverse(), order.slice(k+1));
                        if (cost(cand) < cost(order) - 1e-9) { order = cand; improved = true; }
                    }
            }
            return order;
        }
        function rtGeocode(q) {
            return new Promise(res => {
                const g = new kakao.maps.services.Geocoder();
                g.addressSearch(q, (r, st) => {
                    if (st === kakao.maps.services.Status.OK && r.length)
                        return res({lat: +r[0].y, lng: +r[0].x});
                    const p = new kakao.maps.services.Places();
                    p.keywordSearch(q, (r2, st2) => {
                        if (st2 === kakao.maps.services.Status.OK && r2.length)
                            return res({lat: +r2[0].y, lng: +r2[0].x});
                        res(null);
                    });
                });
            });
        }
        let rtMap = null, rtObjs = [];
        async function rtRun() {
            if (!sdkLoaded) { document.getElementById('rt-msg').textContent = '지도 모듈 로딩 중 — 잠시 후 다시 눌러주세요.'; return; }
            const lines = document.getElementById('rt-input').value.split('\n').map(s => s.trim()).filter(Boolean);
            const msg = document.getElementById('rt-msg');
            if (lines.length < 2) { msg.textContent = '출발지 포함 2곳 이상 입력하세요.'; return; }
            if (lines.length > 40) { msg.textContent = '한 번에 40곳까지 지원합니다.'; return; }
            msg.textContent = '주소 변환 중...';
            const pts = [];
            for (const q of lines) {
                const c = await rtGeocode(q);
                if (!c) { msg.textContent = '주소를 찾지 못했습니다: ' + q; return; }
                pts.push({...c, name: q});
            }
            const roundTrip = document.getElementById('rt-round').checked;
            const order = optimizeRoute(pts, roundTrip);
            // 지도 렌더
            const box = document.getElementById('rt-map'); box.style.display = 'block';
            if (!rtMap) rtMap = new kakao.maps.Map(box, {center: new kakao.maps.LatLng(37.6, 127.06), level: 7});
            rtObjs.forEach(o => o.setMap(null)); rtObjs = [];
            rtMap.relayout();
            const bounds = new kakao.maps.LatLngBounds(), path = [];
            order.forEach((idx, i) => {
                const p = pts[idx], pos = new kakao.maps.LatLng(p.lat, p.lng);
                path.push(pos); bounds.extend(pos);
                const ov = new kakao.maps.CustomOverlay({position: pos, yAnchor: 0.5,
                    content: '<div class="rt-pin' + (i === 0 ? ' rt-pin-start' : '') + '">' + (i === 0 ? '출발' : i) + '</div>'});
                ov.setMap(rtMap); rtObjs.push(ov);
            });
            if (roundTrip) path.push(path[0]);
            const line = new kakao.maps.Polyline({path: path, strokeWeight: 3, strokeColor: '#2f6bd8', strokeOpacity: 0.85, strokeStyle: 'shortdash'});
            line.setMap(rtMap); rtObjs.push(line);
            rtMap.setBounds(bounds, 40);
            // 목록 렌더
            const list = document.getElementById('rt-list'); list.innerHTML = '';
            let total = 0;
            order.forEach((idx, i) => {
                let leg = '';
                if (i > 0) { const d = haver(pts[order[i-1]], pts[idx]); total += d; leg = ' <span class="rt-km">+' + d.toFixed(1) + 'km</span>'; }
                const li = document.createElement('li');
                li.innerHTML = (i === 0 ? '<b>[출발]</b> ' : '') + pts[idx].name + leg;
                list.appendChild(li);
            });
            if (roundTrip) {
                const d = haver(pts[order[order.length-1]], pts[order[0]]); total += d;
                const li = document.createElement('li');
                li.innerHTML = '<b>[복귀]</b> ' + pts[order[0]].name + ' <span class="rt-km">+' + d.toFixed(1) + 'km</span>';
                list.appendChild(li);
            }
            msg.textContent = '✅ 총 ' + (order.length - 1 + (roundTrip ? 1 : 0)) + '개 구간 · 직선거리 합계 약 ' + total.toFixed(1) + 'km';
        }
        document.getElementById('rt-btn') && document.getElementById('rt-btn').addEventListener('click', rtRun);
"""


# ──────────────────────────────────────────────
# 실행
# ──────────────────────────────────────────────
def main():
    if not NAVER_CLIENT_ID or not NAVER_CLIENT_SECRET:
        raise SystemExit("환경변수 NAVER_CLIENT_ID / NAVER_CLIENT_SECRET 이 설정되지 않았습니다.")

    import urllib3
    urllib3.disable_warnings()  # 일부 구청 서버의 구형 SSL 인증서 경고 무시

    today = datetime.now(KST).replace(tzinfo=None)
    news = collect_news(today)
    deals = collect_deals(today)
    apply_geocoding(deals)
    progress, prog_asof = load_progress()
    land = collect_land(today)
    toheo = collect_toheo(today)

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write(build_html(news, deals, progress, prog_asof, land, toheo, today))
    total_news = sum(len(v) for v in news.values())
    total_deals = sum(len(v) for v in deals.values())
    print(f"\n✅ 생성 완료: {OUTPUT_PATH} (뉴스 {total_news}건 / 실거래 {total_deals}건 / 게시판 바로가기 {len(DISTRICTS)}개)")


if __name__ == "__main__":
    main()
